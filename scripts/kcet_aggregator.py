import math
import os

from openpyxl.styles import Border, Side, PatternFill, Font, Alignment

import pandas as pd
import re

from openpyxl.reader.excel import load_workbook
from pandas.core.computation.ops import isnumeric

def process_excel_data(file_path):
    """
    Loads an Excel file, extracts course codes and names, and processes branch names.

    Args:
        file_path (str): The path to the Excel file.

    Returns:
        list: A list of dictionaries, where each dictionary contains the
              original branch name and the processed branch name (if applicable).
              Returns None if the file is not found or other error occurs.
    """
    try:
        # Read the Excel file into a pandas DataFrame
        df = pd.read_excel(file_path)
        # Check if the DataFrame is empty
        if df.empty:
            print("Error: The Excel file is empty.")
            return None

        # Extract course codes and names, handling potential errors
        if 'COURSE CODE' not in df or 'COURSE DETAIL' not in df or 'INTERESTED' not in df:
            print("Error: The Excel file must contain columns named 'COURSE CODE' and 'COURSE DETAIL' and 'INTERESTED' columns.")
            raise Exception("Error: The Excel file must contain columns named 'COURSE CODE' and 'COURSE DETAIL' and 'INTERESTED' columns.")

        # Filter the DataFrame to only include rows where INTERESTED is 'Y'
        interested_courses_df = df[df['INTERESTED'] == 'Y']

        # Convert to list of dictionaries, handling errors during conversion
        try:
            course_data = interested_courses_df[['COURSE CODE', 'COURSE DETAIL']].to_dict(orient='records')
        except KeyError as e:
            print(f"Error: Column not found: {e}")
            return None
        except Exception as e:
            print(f"Error converting data to dictionary: {e}")
            return None

        # Print the extracted course data
        print("Course Data:")
        for course in course_data:
            print(course)

    except FileNotFoundError:
        print(f"Error: File not found at {file_path}")
        return None
    except Exception as e:
        print(f"An error occurred: {e}")
        return None

    return course_data

def create_dataframe_from_list(data):
    """
    Converts a list of dictionaries to a pandas DataFrame with the specified format.

    Args:
        data (list): A list of dictionaries, where each dictionary represents data for a college.
            Example:
            [
                {"College Code": "E001", "College Name": "Name", "a": "b"},
                {"College Code": "E001", "College Name": "Name", "P": "q"},
                {"College Code": "E001", "College Name": "Name", "l": "m"}
            ]

    Returns:
        pandas.DataFrame: A DataFrame with columns 'College Code', 'College Name', and other keys
        from the dictionaries as columns, with corresponding values in a single row per college.
    """
    if not data:
        return pd.DataFrame()  # Return an empty DataFrame if the input list is empty
    # Extract unique college codes
    college_codes = set(item['College Code'] for item in data)

    # Create a list to store the final data rows
    final_data = []

    for code in college_codes:
        # Filter data for the current college code
        college_data = [item for item in data if item['College Code'] == code]


        # Start with the base data (College Code and Name)
        first_item = college_data[0]
        college_row = {'College Code': first_item['College Code'], 'College Name': first_item['College Name']}

        # Update the dictionary with the other key-value pairs
        college_row_branches = {}
        for item in college_data:
            for key, value in item.items():
                if key not in ['College Code', 'College Name']:
                    college_row_branches[key] = value

        college_row_branches = dict(sorted(college_row_branches.items()))
        college_row.update(college_row_branches)
        final_data.append(college_row)
        # print(final_data)

    # Create the DataFrame
    final_df = pd.DataFrame(final_data)
    return final_df



def extract_and_format_data(course_data, excel_file_path, output_file_path=None):
    """
    Extracts college data from an Excel file with a specific format and converts it to a structured DataFrame,
    with "College Code" and "College Name" columns, followed by branch columns containing the "GM" cutoff rank.

    Args:
        excel_file_path (str): The path to the Excel file.
        output_file_path (str, optional): The path to the output Excel file. Defaults to "formatted_data.xlsx".
    """
    try:
        # Read all sheets from the Excel file
        excel_data = pd.read_excel(excel_file_path, sheet_name=None)

        all_data = []  # List to store processed data from all sheets

        # Iterate through each sheet in the Excel file
        for sheet_name, df in excel_data.items():
            print('processing sheet ', sheet_name)

            data = []
            college_code = None
            college_name = None
            gm_index = -1
            new_section = False
            # Iterate through the rows
            for i in range(len(df)):
                row = df.iloc[i]

                # Check if the row contains college code and name using regex
                if isinstance(row[0], str):
                    match = re.search(r".*?(E\d+)\s+(.*)", row[0])
                    if match:
                        college_code = match.group(1)
                        college_name = match.group(2)
                        new_section = True
                        gm_index = -1
                        continue  # Skip to the next row

                # If college code and name are found, process the table
                if college_code and college_name:
                    # Assuming table headers are in the same row
                    if "BAN" not in college_name.upper() and "BEN" not in college_name.upper():
                        print(f'skip college {college_name}')
                        continue

                    if "KALBURGI" in college_name.upper():
                        print(f'skip college {college_name}')
                        continue
                    if "BANTWAL" in college_name.upper():
                        print(f'skip college {college_name}')
                        continue
                    if "BANGARAPET" in college_name.upper():
                        print(f'skip college {college_name}')
                        continue
                    if "MANGALORE" in college_name.upper():
                        print(f'skip college {college_name}')
                        continue
                    if "RANEBENNUR" in college_name.upper():
                        print(f'skip college {college_name}')
                        continue

                    headers = df.iloc[i].tolist()
                    if gm_index == -1:
                        for j, header in enumerate(headers):
                            if "GM" == header:
                                gm_index = j
                                break
                    branch_gm_data = {'College Code': college_code, 'College Name': college_name}

                    # check if the next row exists before accessing
                    if i + 1 < len(df):
                        branch_row = df.iloc[i + 1].tolist()
                        branch_name = branch_row[0]
                        if isinstance(branch_name, str):
                            interested_course =False
                            for course in course_data:
                                course_code = course['COURSE CODE']
                                course_name = course['COURSE DETAIL']
                                if branch_name.strip().startswith(course_code+ " "):
                                    branch_name = f"{course_code} [{course_name}]"
                                    interested_course = True
                                    break  # Stop after the first match
                            if not interested_course:
                                continue


                            gm_value = branch_row[gm_index]
                            try:
                                float(gm_value)
                            except ValueError:
                                continue
                            gm_value = float(gm_value)
                            if not math.isnan(gm_value):
                                branch_gm_data.update({branch_name: gm_value})
                                data.append(branch_gm_data)

            # Convert the list of dictionaries to a DataFrame for the current sheet
            #formatted_df = create_dataframe_from_list(data)
            all_data.extend(data)
            print('   appended to all data')

        print("ALL SHEETS DONE")
        # Convert the list of dictionaries to a DataFrame
        ##final_df = pd.DataFrame(data)
        final_df = create_dataframe_from_list(all_data)
        print("final_df DONE")
        # Save the final DataFrame to a new Excel file
        #final_df.to_excel(output_file_path, index=False)

        output_file_path = output_file_path or excel_file_path

        with pd.ExcelWriter(output_file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as excel_writer:
            final_df.to_excel(excel_writer, sheet_name='AGGREGATED', index=False)

        print(f"Formatted data saved to {output_file_path}")

        # Load the workbook and reorder sheets
        wb = load_workbook(output_file_path)
        sheet_to_move = wb['AGGREGATED']  # Get the sheet to move
        wb._sheets.remove(sheet_to_move)  # Remove the sheet from its current position
        wb._sheets.insert(0, sheet_to_move)  # Insert at the desired position (0 for the first position)
        wb.save(output_file_path)
        print(f"Formatted data moved as first sheet")

        # Apply formatting
        sheet = wb['AGGREGATED']

        # Define the borders
        border_style = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        default_font = Font(name='Arial', size=10)

        header_font = Font(name='Arial', size=10, bold=True)
        # Define fill
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Light Blue


        # Set column widths for the first two columns
        sheet.column_dimensions['A'].width = 10  # Adjust as needed
        sheet.column_dimensions['B'].width = 90  # Adjust as needed

        # Apply default width and formatting to all columns
        for column_letter in sheet.columns:
            col_dim = sheet.column_dimensions[column_letter[0].column_letter]
            if column_letter[0].column_letter not in ['A', 'B']:
                col_dim.width = 18  # Default width for other columns
            for row_index, cell in enumerate(column_letter):
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                cell.border = border_style
                if row_index == 0:  # Header row
                    cell.font = header_font
                    cell.fill = header_fill
                else:
                    cell.font = default_font

        wb.save(output_file_path)


    except FileNotFoundError:
        print(f"Error: File not found at {excel_file_path}")
    except Exception as e:
        print(f"An error occurred: {e}")


if __name__ == "__main__":
    course_data = process_excel_data("../kcet_config/COURSECODE_ENGGkannada.xlsx")

    # Get all Excel files in the directory
    excel_files = [f for f in os.listdir("../kcet_files") if f.endswith(('.xlsx', '.xls'))]

    for excel_file in excel_files:
        excel_file_path = os.path.join("../kcet_files", excel_file)
        print(f"Processing file: {excel_file_path}")
        extract_and_format_data(course_data, excel_file_path)  # Pass the output file path

    print("All files processed.")
