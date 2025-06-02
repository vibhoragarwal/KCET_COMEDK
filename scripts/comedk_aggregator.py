import os

import pandas as pd
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment

def coalesce_and_clean_merged_columns(df):
    """
    Identifies columns ending with '_x' and '_y' in a DataFrame,
    coalesces them into a single column (preferring _y if available),
    and then drops the original suffixed columns.

    Args:
        df (pd.DataFrame): The DataFrame after a merge operation that resulted
                           in _x and _y suffixes.

    Returns:
        pd.DataFrame: The DataFrame with coalesced and cleaned column names.
    """
    columns_to_process = {}
    for col in df.columns:
        if col.endswith('_x'):
            base_name = col[:-2] # Remove '_x'
            y_col = f"{base_name}_y"
            if y_col in df.columns:
                columns_to_process[base_name] = (col, y_col)

    if not columns_to_process:
        print("No columns found ending with '_x' and '_y' for coalescing.")
        return df # No changes needed

    print(f"Coalescing {len(columns_to_process)} sets of columns:")
    cols_to_drop = []
    for base_name, (x_col, y_col) in columns_to_process.items():
        print(f"  - Coalescing '{x_col}' and '{y_col}' into '{base_name}'")
        # Coalesce: prefer value from _y (right DataFrame), fall back to _x (left DataFrame)
        df[base_name] = df[y_col].fillna(df[x_col])
        cols_to_drop.extend([x_col, y_col]) # Mark suffixed columns for dropping

    df.drop(columns=cols_to_drop, inplace=True)
    print("Coalescing complete and original suffixed columns dropped.")
    return df


def combine_sheets_from_multiple_excels(input_excel_paths, output_excel_path):
    """
    Loads all sheets from a list of Excel files and saves them into a single new Excel file,
    with each original sheet becoming a new tab.

    Args:
        input_excel_paths (list): A list of file paths to the input Excel files.
        output_excel_path (str): The file path for the output Excel file (e.g., "X.xlsx").
    """
    print(f"Starting to combine sheets into '{output_excel_path}'...")

    sheets = []

    # Create an ExcelWriter object to write multiple sheets to a single Excel file
    with pd.ExcelWriter(output_excel_path, engine='xlsxwriter') as writer:
        for excel_file_path in input_excel_paths:
            if not os.path.exists(excel_file_path):
                print(f"  Warning: '{excel_file_path}' not found. Skipping.")
                continue

            try:
                # Read all sheets from the current Excel file
                xls = pd.ExcelFile(excel_file_path)
                sheet_names = xls.sheet_names
                print(f"  Processing '{os.path.basename(excel_file_path)}' with sheets: {sheet_names}")

                for sheet_name in sheet_names:
                    df = pd.read_excel(xls, sheet_name=sheet_name)

                    # Write the DataFrame to a new sheet in the output Excel file
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    print(f"    Added sheet '{sheet_name}' as '{sheet_name}'.")
                    sheets.append(sheet_name)

            except Exception as e:
                print(f"  Error processing '{excel_file_path}': {e}")
            print("-" * 50)

    print(f"Successfully combined all sheets into '{output_excel_path}'.")

    # Load the workbook and reorder sheets
    wb = load_workbook(output_excel_path)

    for sheet_n in sheets:

        # Apply formatting
        sheet = wb[sheet_n]

        # Define the borders
        border_style = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        default_font = Font(name='Arial', size=10)

        header_font = Font(name='Arial', size=10, bold=True)
        # Define fill
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Light Blue

        # Set column widths for the first two columns
        sheet.column_dimensions['A'].width = 10  # Adjust as needed
        sheet.column_dimensions['B'].width = 90  # Adjust as needed

        # Apply default width and formatting to all columns
        for column_letter in sheet.columns:
            col_dim = sheet.column_dimensions[column_letter[0].column_letter]
            if column_letter[0].column_letter not in ['A', 'B']:
                col_dim.width = 18  # Default width for other columns
            for row_index, cell in enumerate(column_letter):
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                cell.border = border_style
                if row_index == 0:  # Header row
                    cell.font = header_font
                    cell.fill = header_fill
                else:
                    cell.font = default_font

    wb.save(output_excel_path)


def process_comedk_data(folder_path, branch_codes_file_name, input_excel_name):
    """
    Processes COMEDK ranking data to extract specific branches for GM category,
    ensuring all interested branches are present in the output with their cut-offs
    or pd.NA if not offered/data missing for a college.

    Args:
        folder_path (str): The path to the folder containing the input Excel files.
        branch_codes_file_name (str): The name of the Excel file containing branch codes
                                      and interest flags (e.g., 'COMEDK_BRANCH_CODES.xlsx').
        input_excel_name (str): input file

    Returns:
        pd.DataFrame: A DataFrame containing the processed data, or None if no data is found.
    """
    try:
        # Construct full paths
        branch_codes_full_path = os.path.join(folder_path, branch_codes_file_name)

        # 1. Read COMEDK_BRANCH_CODES.xlsx to get interested branch codes and their full names
        print(f"Reading branch codes from: {branch_codes_full_path}")
        df_branch_codes = pd.read_excel(branch_codes_full_path)
        interested_branches_df = df_branch_codes[df_branch_codes['Interested'] == 'Y'].copy()

        # Create a list of full branch column names (e.g., '30316-Aeronautical Engineering')
        # for all interested branches. This defines the final desired columns for branches.
        all_potential_interested_branch_columns = [
            f"{row['Branch Code']}-{row['Branch Name']}"
            for index, row in interested_branches_df.iterrows()
        ]

        all_processed_data = []

        # Find the COMEDK_R*.xlsx file(s) in the specified folder
        input_file = os.path.join(folder_path, input_excel_name)

        # Define the base columns that should always be present
        base_columns = ['College Code', 'College Name', 'Seat Category']

        # 2. Process the COMEDK_R*.xlsx file(s)
        print(f"Processing COMEDK ranking file: {input_file}")
        xl = pd.ExcelFile(input_file)
        for sheet_name in xl.sheet_names:
            print(f'  processing sheet {sheet_name}')
            df_sheet = pd.read_excel(xl, sheet_name=sheet_name)
            if 'Seat Type' in df_sheet.columns:
                df_sheet.rename(columns={'Seat Type': 'Seat Category'}, inplace=True)
                print(f"Renamed 'Seat Type' to 'Seat Category' in '{input_file}'.")
            elif 'Seat type' in df_sheet.columns:
                df_sheet.rename(columns={'Seat type': 'Seat Category'}, inplace=True)
                print(f"Renamed 'Seat type' to 'Seat Category' in '{input_file}'.")
            elif 'Seat Category' not in df_sheet.columns:
                print(f"Warning: Neither 'Seat Type' nor 'Seat Category' found in '{input_file}'.")
                # Optional: df_sheet['Seat Category'] = pd.NA # Add an empty column if needed

            # Ensure base columns exist in the current sheet
            if not all(col in df_sheet.columns for col in base_columns):
                missing_cols = [col for col in base_columns if col not in df_sheet.columns]
                print(
                    f"Error: Required base columns {missing_cols} not found in sheet '{sheet_name}' of file '{input_file}'. Skipping this sheet.")
                continue

            # Filter for 'GM' category
            df_gm = df_sheet[df_sheet['Seat Category'] == 'GM'].copy()

            # Create a temporary DataFrame to hold the selected and aligned data for this sheet
            # Initialize with base columns and all potential interested branch columns, filled with pd.NA
            # This ensures all target columns are present from the start for this sheet's data
            df_aligned_sheet = pd.DataFrame()
            # Populate base columns
            for col in base_columns:
                if col in df_gm.columns:
                    df_aligned_sheet[col] = df_gm[col]
                # If a base column is missing, it was already caught by the 'if not all(col in df_sheet.columns...' check

            # Populate interested branch columns, preserving existing data
            for interested_col_name in all_potential_interested_branch_columns:
                # Check if the interested column exists in the current sheet's GM filtered data
                if interested_col_name in df_gm.columns:
                    df_aligned_sheet[interested_col_name] = df_gm[interested_col_name]
            all_processed_data.append(df_aligned_sheet)

        if not all_processed_data:
            print("No 'GM' category data found in any of the input files for the specified branches.")
            return None

            # Initialize the merged_df with the first processed DataFrame
            # Using .copy() to avoid potential SettingWithCopyWarning later
        merged_final_df = all_processed_data[0].copy()
        # print(merged_final_df.to_string())

        # Merge sequentially with subsequent DataFrames
        for i in range(1, len(all_processed_data)):
            df_to_merge = all_processed_data[i].copy()  # Ensure we're merging a copy
            # print(df_to_merge.to_string())
            # Perform the outer merge on base_columns.
            # Since additional columns have already been uniquely renamed (e.g., '_R1', '_R2'),
            # pandas won't need to add its own suffixes and won't find conflicts.
            merged_final_df = pd.merge(merged_final_df, df_to_merge,
                                       on=base_columns,
                                       how='outer'
                                       )

        merged_final_df = coalesce_and_clean_merged_columns(merged_final_df)

        all_columns = merged_final_df.columns.tolist()
        # Separate base and additional columns
        additional_columns = [col for col in all_columns if col not in base_columns]
        # Create the desired column order
        final_column_order = base_columns + additional_columns
        merged_final_df = merged_final_df[final_column_order]
        return merged_final_df

    except FileNotFoundError as e:
        print(f"Error: One of the files not found. Please ensure the folder path and file names are correct. {e}")
        return None
    except KeyError as e:
        print(
            f"Error: Missing expected column. {e}. Please check your Excel headers in '{branch_codes_full_path}' or the COMEDK_R file(s).")
        return None
    except Exception as e:
        print(f"An unexpected error occurred: {e}")
        return None


# --- Main execution ---
if __name__ == "__main__":
    # Define the folder where your input files are located
    input_folder = '../comedk_files'
    # Ensure the input folder exists relative to the script's location
    absolute_input_folder = os.path.abspath(input_folder)

    branch_codes_file_name = 'COMEDK_BRANCH_CODES.xlsx'

    file_data = [{'COMEDK_R1_12_07_2024.xlsx': "COMEDK_R1_12_07_2024_output.xlsx"},
                 {'COMEDK_R2_07_08_2024.xlsx': "COMEDK_R2_07_08_2024_output.xlsx"},
                 {'COMEDK_R3_09_09_2024.xlsx': "COMEDK_R3_09_09_2024_output.xlsx"}
                 ]

    wb_lists = []

    for file_map in file_data:
        # Each 'file_map' is a dictionary with a single key-value pair
        for input_file, output_file in file_map.items():
            print(f"Processing input file: {input_file}")
            print(f"Output will be saved to: {output_file}")

            # --- Your code goes here to process each file ---
            try:

                # Call the processing function
                final_output_df = process_comedk_data(absolute_input_folder, branch_codes_file_name, input_file)

                # 5. Generate output Excel file in the same input folder
                output_full_path = os.path.join(absolute_input_folder, output_file)

                sheet_name_from_input = os.path.splitext(os.path.basename(input_file))[0]
                final_output_df.to_excel(output_full_path, index=False, sheet_name=sheet_name_from_input)

                wb_lists.append(output_full_path)

                print(f"\nSuccessfully generated '{output_full_path}' with the processed data.")
            except FileNotFoundError:
                print(f"Error: The file '{input_file}' was not found. Please check the path.\n")
            except Exception as e:
                print(f"An unexpected error occurred while processing '{input_file}': {e}\n")

    combine_sheets_from_multiple_excels(wb_lists, os.path.join(absolute_input_folder, 'COMEDK_ALL_2024_output.xlsx'))
